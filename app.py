import os
import streamlit as st
import pandas as pd
import gspread
import folium
from streamlit_folium import st_folium
from google.oauth2.service_account import Credentials
from datetime import datetime
from copy import copy
from difflib import SequenceMatcher

import io
import csv
import re
import zipfile
import smtplib
import tempfile
import shutil
from collections import defaultdict
from email.message import EmailMessage
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

st.set_page_config(page_title="Nu Life Admin App", page_icon="📍", layout="wide")

# =========================
# DARK PREMIUM APP STYLING
# =========================
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(180deg, #020617 0%, #071827 100%);
        color: #e2e8f0;
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #020617 0%, #0f2a3d 100%);
    }

    section[data-testid="stSidebar"] * {
        color: #e2e8f0 !important;
    }

    h1, h2, h3, h4 {
        color: #f8fafc !important;
        font-weight: 800;
    }

    p, label, span, div {
        color: inherit;
    }

    div[data-testid="stMetric"] {
        background: #0f172a;
        border: 1px solid #1e293b;
        padding: 18px;
        border-radius: 18px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.35);
    }

    div[data-testid="stMetric"] * {
        color: #f8fafc !important;
    }

    .stButton > button {
        border-radius: 12px;
        border: none;
        background: #22c55e;
        color: #020617;
        font-weight: 800;
        padding: 0.6rem 1rem;
    }

    .stButton > button:hover {
        background: #16a34a;
        color: white;
    }

    div[data-testid="stDataFrame"] {
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 10px 30px rgba(0,0,0,0.35);
    }

    .block-container {
        padding-top: 2rem;
    }

    hr {
        border: none;
        border-top: 1px solid #1e293b;
    }

    .premium-card {
        background: #0f172a;
        border: 1px solid #1e293b;
        border-radius: 18px;
        padding: 18px;
        margin-bottom: 14px;
        box-shadow: 0 10px 30px rgba(0,0,0,0.35);
        color: #e2e8f0;
    }

    .premium-card-title {
        font-size: 22px;
        font-weight: 900;
        color: #f8fafc;
    }

    .premium-card-subtitle {
        font-size: 14px;
        color: #94a3b8;
        margin-top: 4px;
    }

    .premium-card-body {
        margin-top: 10px;
        color: #cbd5e1;
    }
</style>
""", unsafe_allow_html=True)

# =========================
# DARK HEADER
# =========================
st.markdown("""
<div style="
    background: linear-gradient(135deg, #071827 0%, #0f172a 100%);
    border: 1px solid #1e293b;
    border-radius: 22px;
    padding: 22px 28px;
    margin-bottom: 24px;
    box-shadow: 0 10px 40px rgba(0,0,0,0.55);
">
""", unsafe_allow_html=True)

if os.path.exists("logo.png"):
    col1, col2 = st.columns([1, 5])
    with col1:
        st.image("logo.png", width=130)
    with col2:
        st.markdown("""
        <div style="font-size: 34px; font-weight: 900; color: white; padding-top: 10px;">
            Nu Life Admin App
        </div>
        <div style="font-size: 15px; color: #94a3b8; margin-top: 4px;">
            Sales territory mapping, rep profiles, and performance visibility
        </div>
        """, unsafe_allow_html=True)
else:
    st.markdown("""
    <div style="font-size: 34px; font-weight: 900; color: white;">
        Nu Life Admin App
    </div>
    <div style="font-size: 15px; color: #94a3b8; margin-top: 4px;">
        Sales territory mapping, rep profiles, and performance visibility
    </div>
    """, unsafe_allow_html=True)

st.markdown("</div>", unsafe_allow_html=True)

GSHEET_ID = st.secrets["GSHEET_ID"]
APP_PASSWORD = st.secrets["APP_PASSWORD"]

REP_HEADERS = [
    "RepID", "Active", "Manager", "Region", "MarketTerritory", "State", "City",
    "FirstName", "LastName", "FullName", "PhoneNumber", "PersonalEmail",
    "NuLifeEmail", "LinksHandles", "BusinessName", "Address", "Latitude",
    "Longitude", "Notes", "StartDate", "LastUpdated"
]

SALES_HEADERS = [
    "Date", "RepID", "FullName", "MarketTerritory", "State", "Orders",
    "Revenue", "Providers", "TopProduct", "LastOrderDate", "AverageOrderValue"
]

SALES_HISTORY_HEADERS = [
    "ReportPeriod", "Month", "RepName", "Orders", "Sales", "AverageOrder",
    "TopProduct1", "TopProduct2", "TopProduct3", "CancelledOrders",
    "TotalRows", "CancelledRate", "GeneratedAt"
]

def get_gsheet_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scopes
    )
    return gspread.authorize(creds)

@st.cache_data(ttl=300)
def load_reps():
    try:
        gc = get_gsheet_client()
        sheet = gc.open_by_key(GSHEET_ID)
        ws = sheet.worksheet("rep_profiles")
        data = ws.get_all_records()
        df = pd.DataFrame(data)

        for col in REP_HEADERS:
            if col not in df.columns:
                df[col] = ""

        return df[REP_HEADERS]

    except Exception as e:
        st.error("Google Sheets rep_profiles connection failed.")
        st.write("Error type:", type(e).__name__)
        st.write("Error details:", str(e))
        st.stop()

@st.cache_data(ttl=300)
def load_sales():
    try:
        gc = get_gsheet_client()
        sheet = gc.open_by_key(GSHEET_ID)
        ws = sheet.worksheet("rep_sales")
        data = ws.get_all_records()
        df = pd.DataFrame(data)

        for col in SALES_HEADERS:
            if col not in df.columns:
                df[col] = ""

        return df[SALES_HEADERS]

    except Exception:
        return pd.DataFrame(columns=SALES_HEADERS)

def get_or_create_worksheet(sheet, worksheet_name, headers):
    try:
        ws = sheet.worksheet(worksheet_name)
    except Exception:
        ws = sheet.add_worksheet(title=worksheet_name, rows=1000, cols=max(len(headers), 10))
        ws.update([headers])
    return ws

@st.cache_data(ttl=300)
def load_sales_history():
    try:
        gc = get_gsheet_client()
        sheet = gc.open_by_key(GSHEET_ID)
        ws = get_or_create_worksheet(sheet, "rep_sales_history", SALES_HISTORY_HEADERS)
        data = ws.get_all_records()
        df = pd.DataFrame(data)

        for col in SALES_HISTORY_HEADERS:
            if col not in df.columns:
                df[col] = ""

        return df[SALES_HISTORY_HEADERS]

    except Exception:
        return pd.DataFrame(columns=SALES_HISTORY_HEADERS)

def save_sales_history(df):
    try:
        gc = get_gsheet_client()
        sheet = gc.open_by_key(GSHEET_ID)
        ws = get_or_create_worksheet(sheet, "rep_sales_history", SALES_HISTORY_HEADERS)

        clean_df = df.copy()
        for col in SALES_HISTORY_HEADERS:
            if col not in clean_df.columns:
                clean_df[col] = ""

        clean_df = clean_df[SALES_HISTORY_HEADERS].fillna("")
        ws.clear()
        ws.update([SALES_HISTORY_HEADERS] + clean_df.astype(str).values.tolist())
        st.cache_data.clear()
        return True

    except Exception as e:
        st.error("Could not save rep_sales_history to Google Sheets.")
        st.write("Error type:", type(e).__name__)
        st.write("Error details:", str(e))
        return False

def save_reps(df):
    try:
        gc = get_gsheet_client()
        sheet = gc.open_by_key(GSHEET_ID)
        ws = sheet.worksheet("rep_profiles")

        clean_df = df.copy()
        for col in REP_HEADERS:
            if col not in clean_df.columns:
                clean_df[col] = ""

        clean_df = clean_df[REP_HEADERS].fillna("")
        ws.clear()
        ws.update([REP_HEADERS] + clean_df.astype(str).values.tolist())
        st.cache_data.clear()
        return True

    except Exception as e:
        st.error("Could not save data to Google Sheets.")
        st.write("Error type:", type(e).__name__)
        st.write("Error details:", str(e))
        return False

def stable_offset(index):
    offsets = [
        (0.0000, 0.0000),
        (0.0080, 0.0080),
        (-0.0080, -0.0080),
        (0.0080, -0.0080),
        (-0.0080, 0.0080),
    ]
    return offsets[index % len(offsets)]

def clean_sales_df(sales_df):
    df = sales_df.copy()

    if df.empty:
        return df

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["LastOrderDate"] = pd.to_datetime(df["LastOrderDate"], errors="coerce")
    df["Orders"] = pd.to_numeric(df["Orders"], errors="coerce").fillna(0)
    df["Revenue"] = pd.to_numeric(df["Revenue"], errors="coerce").fillna(0)
    df["Providers"] = pd.to_numeric(df["Providers"], errors="coerce").fillna(0)
    df["AverageOrderValue"] = pd.to_numeric(df["AverageOrderValue"], errors="coerce").fillna(0)

    return df

def generate_next_rep_id(existing_df):
    existing_ids = existing_df["RepID"].dropna().astype(str).tolist()
    numbers = []

    for rep_id in existing_ids:
        if rep_id.startswith("REP-"):
            try:
                numbers.append(int(rep_id.replace("REP-", "")))
            except:
                pass

    next_number = max(numbers) + 1 if numbers else 1
    return f"REP-{next_number:03d}"


# =========================
# REPORTING ENGINE
# =========================
REPORTS_PARENT_FOLDER_ID = st.secrets.get("REPORTS_PARENT_FOLDER_ID", "")
SENDER_EMAIL = st.secrets.get("SENDER_EMAIL", "")
SENDER_APP_PASSWORD = st.secrets.get("SENDER_APP_PASSWORD", "")

COMMISSION_LEVELS = {
    1: {"name": 20, "pct": 21, "comm": 22},  # U-W
    2: {"name": 23, "pct": 24, "comm": 25},  # X-Z
    3: {"name": 26, "pct": 27, "comm": 28},  # AA-AC
    4: {"name": 29, "pct": 30, "comm": 31},  # AD-AF
}

ORDER_STATUS_IDX = 17
SUBTOTAL_IDX = 11
CUSTOMER_EMAIL_IDX = 5
TRUE_COST_IDX = 44

JOEL_NELSON_RATE = 0.025
HOWARD_FINDER_RATE = 0.05
HOWARD_CLIENT_EMAILS = {
    "shannon@innerglowstudiofl.com",
    "roxybarberap@gmail.com",
}

def report_clean(value):
    return str(value or "").strip()

def report_norm(value):
    return report_clean(value).lower()

def normalize_name(value):
    """
    Strong name normalizer for matching commission report names to rep_profiles.
    Handles case, extra spaces, punctuation, middle initials, suffixes, apostrophes,
    and common first-name variants.
    """
    value = report_clean(value).lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()

    suffixes = {"jr", "sr", "ii", "iii", "iv", "md", "do", "phd"}
    parts = [p for p in value.split() if p not in suffixes]

    nickname_map = {
        "steven": "steve",
        "stephen": "steve",
        "matthew": "matt",
        "michael": "mike",
        "robert": "bob",
        "william": "bill",
        "andrew": "andrew",
        "andrrew": "andrew",
        "rebecca": "becca",
        "timothy": "tim",
        "ronald": "ron",
        "christopher": "chris",
        "joseph": "joe",
        "daniel": "dan",
        "david": "dave",
        "james": "jim",
        "kenneth": "ken",
    }

    if parts:
        parts[0] = nickname_map.get(parts[0], parts[0])

    return " ".join(parts)

def name_tokens(value):
    normalized = normalize_name(value)
    return [p for p in normalized.split() if p]

def name_similarity(a, b):
    return SequenceMatcher(None, normalize_name(a), normalize_name(b)).ratio()

def report_money(value):
    s = report_clean(value).replace("$", "").replace(",", "").replace(" ", "")
    if not s:
        return 0.0
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except Exception:
        return 0.0

def safe_filename(value):
    return re.sub(r"[^A-Za-z0-9._ -]+", "", report_clean(value)).strip() or "Unknown"

def is_cancelled_order(row):
    return "cancel" in report_norm(row[ORDER_STATUS_IDX] if len(row) > ORDER_STATUS_IDX else "")

def parse_report_period(title):
    text = report_clean(title)
    text = text.replace("Nu Life Essentials Rep ", "")
    text = text.replace("Nu Life Essentials ", "")
    text = text.replace("Nu Life ", "")
    return text or "Commission Runs"

def clean_folder_name(title):
    text = report_clean(title)
    text = text.replace("Nu Life Essentials Rep ", "Nu Life ")
    text = text.replace("Nu Life Essentials ", "Nu Life ")
    if "Commission" not in text:
        text = "Nu Life Commission Runs"
    return text

def get_drive_service():
    scopes = [
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets",
    ]
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scopes
    )
    return build("drive", "v3", credentials=creds)

def create_drive_folder(service, name, parent_id):
    metadata = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
    }
    if parent_id:
        metadata["parents"] = [parent_id]
    folder = service.files().create(body=metadata, fields="id, webViewLink").execute()
    return folder["id"], folder.get("webViewLink", "")

def upload_file_to_drive(service, local_path, folder_id, drive_name=None):
    drive_name = drive_name or os.path.basename(local_path)
    metadata = {"name": drive_name, "parents": [folder_id]}
    media = MediaFileUpload(local_path, resumable=False)
    uploaded = service.files().create(
        body=metadata,
        media_body=media,
        fields="id, webViewLink"
    ).execute()
    return uploaded

def keep_cols_for_level(headers, level):
    drop = {TRUE_COST_IDX}
    for lvl in range(1, level):
        start = COMMISSION_LEVELS[lvl]["name"]
        drop.update([start, start + 1, start + 2])
    return [i for i in range(len(headers)) if i not in drop]

def apply_widths(ws):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = max(len(str(cell.value or "")) for cell in ws[letter])
        ws.column_dimensions[letter].width = min(max_len + 3, 45)


def should_force_alex_20(rep_name, level):
    return report_norm(rep_name) == "alex bethel" and int(level) == 1


def force_alex_20_on_display_row(display_row):
    """
    Alex Bethel should be paid 20% on his direct Level 1 commission report.
    After upper columns are removed, Alex's percent is column V and commission is column W.
    Subtotal is column L.
    """
    percent_idx = 21  # V, zero-based after no upper-level removal for Level 1
    commission_idx = 22  # W
    subtotal_idx = SUBTOTAL_IDX  # L

    if len(display_row) > commission_idx:
        display_row[percent_idx] = "20%"
        display_row[commission_idx] = round(report_money(display_row[subtotal_idx]) * 0.20, 2)

    return display_row


def write_report_sheet(ws, rep_name, level, rows, headers, period):
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    red = PatternFill("solid", fgColor="FFC7CE")

    keep = keep_cols_for_level(headers, level)
    kept_headers = [headers[i] for i in keep]
    generated_comm_col = 23  # W after upper rep columns are removed

    ws["A1"] = f"{rep_name} {period}"
    ws["A1"].font = Font(size=16, bold=True, color="000000")
    ws["A1"].alignment = Alignment(horizontal="left")

    for c, h in enumerate(kept_headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    prepared = []
    for r in rows:
        r = r + [""] * (len(headers) - len(r))
        display_row = [r[i] for i in keep]

        if should_force_alex_20(rep_name, level):
            display_row = force_alex_20_on_display_row(display_row)

        prepared.append(display_row)

    prepared.sort(key=lambda row: (
        report_norm(row[17]) if len(row) > 17 else "",
        report_norm(row[23]) if len(row) > 23 else "",
        report_norm(row[4]) if len(row) > 4 else "",
    ))

    total_due = 0.0
    cancelled_count = 0

    for rr, row in enumerate(prepared, 3):
        for cc, val in enumerate(row, 1):
            ws.cell(rr, cc, val)

        if "cancel" in report_norm(ws.cell(rr, 18).value):
            cancelled_count += 1
            for cc in range(1, len(kept_headers) + 1):
                ws.cell(rr, cc).fill = red
                ws.cell(rr, cc).font = Font(color="000000")
            ws.cell(rr, generated_comm_col).value = 0
            payout = 0.0
        else:
            payout = report_money(ws.cell(rr, generated_comm_col).value)

        total_due += payout

    last_data = ws.max_row
    due_row = last_data + 3
    ws[f"F{due_row}"] = "Due"
    ws[f"G{due_row}"] = round(total_due, 2)
    ws[f"F{due_row}"].font = Font(bold=True)
    ws[f"G{due_row}"].font = Font(bold=True)
    ws[f"F{due_row}"].fill = light
    ws[f"G{due_row}"].fill = light
    ws[f"G{due_row}"].number_format = "$#,##0.00"

    for col in ["L", "M", "N", "O", "P", "W", "Z", "AC", "AF"]:
        try:
            for cell in ws[col][2:last_data]:
                cell.number_format = "$#,##0.00"
        except Exception:
            pass

    ws.freeze_panes = "A3"
    apply_widths(ws)
    return total_due, cancelled_count, len(prepared)

def make_cleaned_raw_workbook(title, headers, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Raw Report"
    blue = PatternFill("solid", fgColor="1F4E78")
    red = PatternFill("solid", fgColor="FFC7CE")

    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="000000")
    ws["A1"].alignment = Alignment(horizontal="left")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(2, c, h)
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")

    commission_cols = [
        COMMISSION_LEVELS[1]["comm"],
        COMMISSION_LEVELS[2]["comm"],
        COMMISSION_LEVELS[3]["comm"],
        COMMISSION_LEVELS[4]["comm"],
    ]

    for rr, row in enumerate(data, 3):
        row = row + [""] * (len(headers) - len(row))
        cancelled = is_cancelled_order(row)

        # Force Alex Bethel direct commission from 19% to 20%.
        if not cancelled and report_norm(row[COMMISSION_LEVELS[1]["name"]]) == "alex bethel":
            row[COMMISSION_LEVELS[1]["pct"]] = "20%"
            row[COMMISSION_LEVELS[1]["comm"]] = round(report_money(row[SUBTOTAL_IDX]) * 0.20, 2)

        for cc, val in enumerate(row, 1):
            raw_idx = cc - 1
            if cancelled and raw_idx in commission_cols:
                val = 0
            cell = ws.cell(rr, cc, val)
            if cancelled:
                cell.fill = red
                cell.font = Font(color="000000")

    ws.freeze_panes = "A3"
    apply_widths(ws)
    return wb

def calculate_alex_override(data):
    # Alex override: 1.5% of subtotal revenue that is NOT Dean-only direct and NOT Alex/downline business.
    qualifying_revenue = 0.0
    detail_rows = []

    for row in data:
        row = row + [""] * 60
        if is_cancelled_order(row):
            continue

        rep1 = report_norm(row[COMMISSION_LEVELS[1]["name"]])
        rep2 = report_norm(row[COMMISSION_LEVELS[2]["name"]])
        rep3 = report_norm(row[COMMISSION_LEVELS[3]["name"]])
        rep4 = report_norm(row[COMMISSION_LEVELS[4]["name"]])

        dean_direct = rep1 == "dean baker" and not rep2 and not rep3 and not rep4
        alex_business = "alex bethel" in [rep1, rep2, rep3, rep4]

        if dean_direct or alex_business:
            continue

        qualifying_revenue += report_money(row[SUBTOTAL_IDX])
        detail_rows.append(row)

    return qualifying_revenue, round(qualifying_revenue * 0.015, 2), detail_rows

def add_alex_summary(wb, regular_due, override_revenue, override_due):
    ws = wb.create_sheet("Alex Total Summary", 0)
    ws["A1"] = "Alex Bethel Total Pay Summary"
    ws["A1"].font = Font(size=16, bold=True)
    rows = [
        ("Regular Commission Due", regular_due),
        ("1.5% Override Qualifying Revenue", override_revenue),
        ("1.5% Override Due", override_due),
        ("Total Due", regular_due + override_due),
    ]
    for i, (label, value) in enumerate(rows, 3):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = round(value, 2)
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"].number_format = "$#,##0.00"
    ws["A6"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["B6"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["B6"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18

def add_override_detail_sheet(wb, sheet_name, title, headers, rows, rate, subtotal_idx=SUBTOTAL_IDX):
    ws = wb.create_sheet(sheet_name)
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")
    red = PatternFill("solid", fgColor="FFC7CE")

    keep = [i for i in range(len(headers)) if i != TRUE_COST_IDX]
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")

    for c, h in enumerate([headers[i] for i in keep], 1):
        cell = ws.cell(2, c, h)
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")

    qualifying_revenue = 0.0
    for rr, row in enumerate(rows, 3):
        row = row + [""] * (len(headers) - len(row))
        cancelled = is_cancelled_order(row)
        if not cancelled:
            qualifying_revenue += report_money(row[subtotal_idx])
        for cc, idx in enumerate(keep, 1):
            val = row[idx]
            cell = ws.cell(rr, cc, val)
            if cancelled:
                cell.fill = red
                cell.font = Font(color="000000")

    last_data = ws.max_row
    summary = last_data + 3
    due = round(qualifying_revenue * rate, 2)
    ws[f"F{summary}"] = "Qualifying Revenue"
    ws[f"G{summary}"] = round(qualifying_revenue, 2)
    ws[f"F{summary+1}"] = "Override %"
    ws[f"G{summary+1}"] = rate
    ws[f"F{summary+2}"] = "Due"
    ws[f"G{summary+2}"] = due

    for r in range(summary, summary + 3):
        ws[f"F{r}"].font = Font(bold=True)
        ws[f"G{r}"].font = Font(bold=True)
        ws[f"F{r}"].fill = light
        ws[f"G{r}"].fill = light

    ws[f"G{summary}"].number_format = "$#,##0.00"
    ws[f"G{summary+1}"].number_format = "0.00%"
    ws[f"G{summary+2}"].number_format = "$#,##0.00"
    ws.freeze_panes = "A3"
    apply_widths(ws)
    return qualifying_revenue, due


def find_product_column(headers):
    """
    Finds the most likely product/item column in the raw commission report.
    """
    preferred_terms = [
        "product", "product name", "item", "item name", "sku",
        "description", "line item", "top product", "ordered product"
    ]

    normalized_headers = [report_norm(h) for h in headers]

    for term in preferred_terms:
        for idx, header in enumerate(normalized_headers):
            if term in header:
                return idx

    return None

def unique_non_cancelled_orders(rows, headers):
    """
    De-duplicates rows by invoice/order number if column B exists.
    Cancelled orders are excluded from sales charts.
    """
    seen = set()
    cleaned = []

    for row in rows:
        row = row + [""] * (len(headers) - len(row))

        if is_cancelled_order(row):
            continue

        invoice = report_clean(row[1] if len(row) > 1 else "")
        key = invoice if invoice else "|".join(str(x) for x in row[:10])

        if key in seen:
            continue

        seen.add(key)
        cleaned.append(row)

    return cleaned


def extract_month_from_period(period):
    text = report_clean(period)
    match = re.search(r"(\d{1,2})\.(\d{1,2}).*?(\d{4})", text)

    if not match:
        return text

    month_num = int(match.group(1))
    year = int(match.group(3))

    month_names = [
        "", "January", "February", "March", "April", "May", "June",
        "July", "August", "September", "October", "November", "December"
    ]

    if 1 <= month_num <= 12:
        return f"{month_names[month_num]} {year}"

    return text


def period_sort_key(period):
    text = report_clean(period)
    match = re.search(r"(\d{1,2})\.(\d{1,2}).*?(\d{4})", text)

    if not match:
        return (9999, 99, 99, text)

    month = int(match.group(1))
    day = int(match.group(2))
    year = int(match.group(3))

    return (year, month, day, text)


def normalize_history_df(history_df):
    if history_df is None or history_df.empty:
        return pd.DataFrame(columns=SALES_HISTORY_HEADERS)

    df = history_df.copy()

    for col in SALES_HISTORY_HEADERS:
        if col not in df.columns:
            df[col] = ""

    df = df[SALES_HISTORY_HEADERS].copy()

    numeric_cols = [
        "Orders",
        "Sales",
        "AverageOrder",
        "CancelledOrders",
        "TotalRows",
        "CancelledRate",
    ]

    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    return df


def upsert_sales_history(existing_history_df, new_history_rows):
    history_df = normalize_history_df(existing_history_df)
    new_df = pd.DataFrame(new_history_rows)

    if new_df.empty:
        return history_df

    for col in SALES_HISTORY_HEADERS:
        if col not in new_df.columns:
            new_df[col] = ""

    new_df = new_df[SALES_HISTORY_HEADERS].copy()

    numeric_cols = [
        "Orders",
        "Sales",
        "AverageOrder",
        "CancelledOrders",
        "TotalRows",
        "CancelledRate",
    ]

    for col in numeric_cols:
        new_df[col] = pd.to_numeric(new_df[col], errors="coerce").fillna(0)

    if not history_df.empty:
        new_keys = set(
            zip(
                new_df["RepName"].astype(str),
                new_df["ReportPeriod"].astype(str),
            )
        )

        history_df = history_df[
            ~history_df.apply(
                lambda row: (
                    str(row["RepName"]),
                    str(row["ReportPeriod"])
                ) in new_keys,
                axis=1
            )
        ]

    combined = pd.concat([history_df, new_df], ignore_index=True)
    combined["_sort"] = combined["ReportPeriod"].apply(period_sort_key)
    combined = combined.sort_values(["RepName", "_sort"]).drop(columns=["_sort"])

    return combined[SALES_HISTORY_HEADERS]


def create_sales_history_row(rep_name, period, metrics, generated_at):
    top_products = metrics.get("top_products", [])
    top_names = [name for name, count in top_products]

    total_rows = int(metrics.get("total_rows", 0))
    cancelled_orders = int(metrics.get("cancelled_orders", 0))
    cancelled_rate = (cancelled_orders / total_rows) if total_rows else 0

    return {
        "ReportPeriod": period,
        "Month": extract_month_from_period(period),
        "RepName": rep_name,
        "Orders": int(metrics.get("orders", 0)),
        "Sales": round(float(metrics.get("sales", 0)), 2),
        "AverageOrder": round(float(metrics.get("average_order", 0)), 2),
        "TopProduct1": top_names[0] if len(top_names) > 0 else "",
        "TopProduct2": top_names[1] if len(top_names) > 1 else "",
        "TopProduct3": top_names[2] if len(top_names) > 2 else "",
        "CancelledOrders": cancelled_orders,
        "TotalRows": total_rows,
        "CancelledRate": round(cancelled_rate, 4),
        "GeneratedAt": generated_at,
    }


def build_analytics_tables(history_df, current_period):
    df = normalize_history_df(history_df)

    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    current_df = df[
        df["ReportPeriod"].astype(str) == str(current_period)
    ].copy()

    if current_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    leaderboard = current_df.sort_values("Sales", ascending=False)[
        [
            "RepName",
            "Orders",
            "Sales",
            "AverageOrder",
            "CancelledOrders",
            "CancelledRate",
        ]
    ].copy()

    growth_rows = []

    for rep_name, rep_df in df.groupby("RepName"):
        rep_df = rep_df.copy()
        rep_df["_sort"] = rep_df["ReportPeriod"].apply(period_sort_key)
        rep_df = rep_df.sort_values("_sort")

        current_rows = rep_df[
            rep_df["ReportPeriod"].astype(str) == str(current_period)
        ]

        if current_rows.empty:
            continue

        current_sales = float(current_rows.iloc[-1]["Sales"])
        prior_rows = rep_df[
            rep_df["ReportPeriod"].astype(str) != str(current_period)
        ]

        if prior_rows.empty:
            prior_sales = 0.0
            growth_amount = current_sales
            growth_percent = 0.0
        else:
            prior_sales = float(prior_rows.iloc[-1]["Sales"])
            growth_amount = current_sales - prior_sales
            growth_percent = (growth_amount / prior_sales) if prior_sales else 0.0

        growth_rows.append({
            "RepName": rep_name,
            "PreviousSales": round(prior_sales, 2),
            "CurrentSales": round(current_sales, 2),
            "GrowthAmount": round(growth_amount, 2),
            "GrowthPercent": round(growth_percent, 4),
        })

    fastest_growth = pd.DataFrame(growth_rows)

    if not fastest_growth.empty:
        fastest_growth = fastest_growth.sort_values(
            ["GrowthAmount", "GrowthPercent"],
            ascending=False
        )

    cancelled = current_df.sort_values("CancelledRate", ascending=False)[
        ["RepName", "CancelledOrders", "TotalRows", "CancelledRate"]
    ].copy()

    return leaderboard, fastest_growth, cancelled


def create_sales_analytics_workbook(history_df, current_period):
    leaderboard, fastest_growth, cancelled = build_analytics_tables(
        history_df,
        current_period
    )

    wb = Workbook()
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")

    def write_df(sheet, title, df):
        sheet["A1"] = title
        sheet["A1"].font = Font(size=16, bold=True)
        sheet["A1"].alignment = Alignment(horizontal="left")
        sheet["A1"].fill = light

        if df.empty:
            sheet["A3"] = "No data available yet."
            apply_widths(sheet)
            return

        for c, col in enumerate(df.columns, 1):
            cell = sheet.cell(3, c, col)
            cell.fill = blue
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")

        for r_idx, row in enumerate(df.itertuples(index=False), 4):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(r_idx, c_idx, value)

        for row_cells in sheet.iter_rows(min_row=4):
            for cell in row_cells:
                header = str(sheet.cell(3, cell.column).value)
                if isinstance(cell.value, float):
                    if "Rate" in header or "Percent" in header:
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = "$#,##0.00"

        apply_widths(sheet)

    ws = wb.active
    ws.title = "Leaderboard"
    write_df(ws, f"Leaderboard - {current_period}", leaderboard)

    ws2 = wb.create_sheet("Fastest Growing")
    write_df(ws2, f"Fastest Growing Reps - {current_period}", fastest_growth)

    ws3 = wb.create_sheet("Cancelled Rate")
    write_df(ws3, f"Cancelled Order % - {current_period}", cancelled)

    return wb



def add_sales_insight_tabs(wb, rep_name, source_rows, headers, period, history_df=None):
    """
    Adds Monthly Summary, Product Summary, and Charts tabs to each rep workbook.
    Charts include historical periods from rep_sales_history plus the current report run.
    """
    orders = unique_non_cancelled_orders(source_rows, headers)
    total_source_rows = len(source_rows)
    cancelled_orders = sum(1 for row in source_rows if is_cancelled_order(row))
    total_orders = len(orders)
    total_sales = sum(report_money(row[SUBTOTAL_IDX]) for row in orders)
    avg_order = total_sales / total_orders if total_orders else 0

    product_idx = find_product_column(headers)
    product_counts = defaultdict(int)
    if product_idx is not None:
        for row in orders:
            product = report_clean(row[product_idx] if product_idx < len(row) else "")
            if product:
                product_counts[product] += 1
    top_products = sorted(product_counts.items(), key=lambda item: item[1], reverse=True)[:3]

    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")

    current_history_row = create_sales_history_row(
        rep_name,
        period,
        {
            "orders": total_orders,
            "sales": total_sales,
            "average_order": avg_order,
            "top_products": top_products,
            "cancelled_orders": cancelled_orders,
            "total_rows": total_source_rows,
        },
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    rep_history = upsert_sales_history(normalize_history_df(history_df), [current_history_row])
    rep_history = rep_history[rep_history["RepName"].astype(str).str.lower() == report_norm(rep_name)].copy()
    rep_history["_sort"] = rep_history["ReportPeriod"].apply(period_sort_key)
    rep_history = rep_history.sort_values("_sort").drop(columns=["_sort"])

    # Monthly Summary
    ws = wb.create_sheet("Monthly Summary")
    ws["A1"] = f"{rep_name} Sales Summary"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")
    for cell in ws["A1:G1"][0]:
        cell.fill = light

    summary_headers = ["Period", "Total Orders", "Total Sales", "Average Order"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws.cell(3, c, h)
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for r_idx, (_, hist_row) in enumerate(rep_history.iterrows(), 4):
        ws.cell(r_idx, 1, hist_row["Month"] or hist_row["ReportPeriod"])
        ws.cell(r_idx, 2, int(hist_row["Orders"]))
        ws.cell(r_idx, 3, round(float(hist_row["Sales"]), 2))
        ws.cell(r_idx, 4, round(float(hist_row["AverageOrder"]), 2))
        ws.cell(r_idx, 3).number_format = "$#,##0.00"
        ws.cell(r_idx, 4).number_format = "$#,##0.00"

    ws["F3"] = "Metric"
    ws["G3"] = "Current Value"
    ws["F3"].fill = blue
    ws["G3"].fill = blue
    ws["F3"].font = Font(bold=True, color="FFFFFF")
    ws["G3"].font = Font(bold=True, color="FFFFFF")
    metrics = [
        ("Current Orders", total_orders),
        ("Current Sales", round(total_sales, 2)),
        ("Current Average Order", round(avg_order, 2)),
        ("Cancelled Order %", (cancelled_orders / total_source_rows) if total_source_rows else 0),
    ]
    for r, (metric, value) in enumerate(metrics, 4):
        ws[f"F{r}"] = metric
        ws[f"G{r}"] = value
        ws[f"F{r}"].font = Font(bold=True)
        if "Order %" in metric:
            ws[f"G{r}"].number_format = "0.00%"
        elif metric != "Current Orders":
            ws[f"G{r}"].number_format = "$#,##0.00"
    apply_widths(ws)

    # Product Summary
    ps = wb.create_sheet("Product Summary")
    ps["A1"] = f"{rep_name} Top Products"
    ps["A1"].font = Font(size=16, bold=True)
    for c, h in enumerate(["Rank", "Product", "Orders"], 1):
        cell = ps.cell(3, c, h)
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    if top_products:
        for r, (product, count) in enumerate(top_products, 4):
            ps.cell(r, 1, r - 3)
            ps.cell(r, 2, product)
            ps.cell(r, 3, count)
    else:
        ps["A4"] = "No product column detected"
        ps["B4"] = "Confirm portal export product column"
        ps["C4"] = 0
    apply_widths(ps)

    # Charts
    charts = wb.create_sheet("Charts")
    charts["A1"] = f"{rep_name} Sales Charts"
    charts["A1"].font = Font(size=16, bold=True)

    charts["A3"] = "Period"
    charts["B3"] = "Total Orders"
    charts["C3"] = "Total Sales"
    charts["D3"] = "Average Order"
    for r_idx, (_, hist_row) in enumerate(rep_history.iterrows(), 4):
        charts.cell(r_idx, 1, hist_row["Month"] or hist_row["ReportPeriod"])
        charts.cell(r_idx, 2, int(hist_row["Orders"]))
        charts.cell(r_idx, 3, round(float(hist_row["Sales"]), 2))
        charts.cell(r_idx, 4, round(float(hist_row["AverageOrder"]), 2))
        charts.cell(r_idx, 3).number_format = "$#,##0.00"
        charts.cell(r_idx, 4).number_format = "$#,##0.00"
    history_last_row = max(4, 3 + len(rep_history))

    charts["F3"] = "Product"
    charts["G3"] = "Orders"
    if top_products:
        for r, (product, count) in enumerate(top_products, 4):
            charts.cell(r, 6, product)
            charts.cell(r, 7, count)
    else:
        charts["F4"] = "No product data"
        charts["G4"] = 0

    for cell in charts["A3:D3"][0] + charts["F3:G3"][0]:
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")

    orders_chart = BarChart()
    orders_chart.title = "Total Orders Over Time"
    orders_chart.y_axis.title = "Orders"
    orders_chart.x_axis.title = "Period"
    orders_chart.add_data(Reference(charts, min_col=2, min_row=3, max_row=history_last_row), titles_from_data=True)
    orders_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=history_last_row))
    orders_chart.height = 8
    orders_chart.width = 14
    charts.add_chart(orders_chart, "A8")

    sales_chart = LineChart()
    sales_chart.title = "Total Sales Over Time ($)"
    sales_chart.y_axis.title = "Sales"
    sales_chart.x_axis.title = "Period"
    sales_chart.add_data(Reference(charts, min_col=3, min_row=3, max_row=history_last_row), titles_from_data=True)
    sales_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=history_last_row))
    sales_chart.height = 8
    sales_chart.width = 14
    charts.add_chart(sales_chart, "I8")

    avg_chart = BarChart()
    avg_chart.title = "Average Order Amount Over Time"
    avg_chart.y_axis.title = "Average Order"
    avg_chart.x_axis.title = "Period"
    avg_chart.add_data(Reference(charts, min_col=4, min_row=3, max_row=history_last_row), titles_from_data=True)
    avg_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=history_last_row))
    avg_chart.height = 8
    avg_chart.width = 14
    charts.add_chart(avg_chart, "A24")

    product_chart = BarChart()
    product_chart.title = "Top 3 Most Ordered Products"
    product_chart.y_axis.title = "Orders"
    product_chart.x_axis.title = "Product"
    product_data_end = max(4, 3 + len(top_products))
    product_chart.add_data(Reference(charts, min_col=7, min_row=3, max_row=product_data_end), titles_from_data=True)
    product_chart.set_categories(Reference(charts, min_col=6, min_row=4, max_row=product_data_end))
    product_chart.height = 8
    product_chart.width = 14
    charts.add_chart(product_chart, "I24")
    apply_widths(charts)

    return {
        "orders": total_orders,
        "sales": round(total_sales, 2),
        "average_order": round(avg_order, 2),
        "top_products": top_products,
        "cancelled_orders": cancelled_orders,
        "total_rows": total_source_rows,
    }

def build_commission_package(uploaded_file, reps_df, sales_history_df=None, send_live=False, test_email=""):
    temp_root = tempfile.mkdtemp(prefix="nulife_reports_")
    source_path = os.path.join(temp_root, uploaded_file.name)
    with open(source_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    with open(source_path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    title = rows[0][0] if rows and rows[0] else "Nu Life Commission Runs"
    headers = rows[1]
    data = rows[2:]
    period = parse_report_period(title)
    folder_name = clean_folder_name(title)

    package_dir = os.path.join(temp_root, folder_name)
    rep_dir = os.path.join(package_dir, "Rep Reports")
    os.makedirs(rep_dir, exist_ok=True)

    shutil.copyfile(source_path, os.path.join(package_dir, uploaded_file.name))

    cleaned_wb = make_cleaned_raw_workbook(title, headers, data)
    cleaned_path = os.path.join(package_dir, "Nu Life Cleaned Raw Commission Report.xlsx")
    cleaned_wb.save(cleaned_path)

    rep_level_rows = defaultdict(lambda: defaultdict(list))
    for row in data:
        row = row + [""] * (len(headers) - len(row))
        for level, idxs in COMMISSION_LEVELS.items():
            rep_name = report_clean(row[idxs["name"]])
            if rep_name:
                rep_level_rows[rep_name][level].append(row)

    alex_override_revenue, alex_override_due, alex_override_rows = calculate_alex_override(data)

    pay_entries = []
    delivery_rows = []
    history_update_rows = []
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sales_history_df = normalize_history_df(sales_history_df)

    for rep_name in sorted(rep_level_rows.keys(), key=lambda x: x.lower()):
        wb = Workbook()
        wb.remove(wb.active)
        regular_due = 0.0
        row_count = 0
        cancelled_count = 0

        for level in sorted(rep_level_rows[rep_name].keys()):
            sheet_name = "Commission Report" if len(rep_level_rows[rep_name]) == 1 else f"Level {level} Report"
            ws = wb.create_sheet(sheet_name)
            due, cancelled, count = write_report_sheet(ws, rep_name, level, rep_level_rows[rep_name][level], headers, period)
            regular_due += due
            cancelled_count += cancelled
            row_count += count

        override_due = 0.0
        override_note = ""

        if report_norm(rep_name) == "alex bethel":
            override_due = alex_override_due
            override_note = "Includes 1.5% override"
            add_alex_summary(wb, regular_due, alex_override_revenue, alex_override_due)
            add_override_detail_sheet(
                wb,
                "Alex 1.5 Override Detail",
                f"Alex 1.5% Override {period}",
                headers,
                alex_override_rows,
                0.015
            )

        total_due = regular_due + override_due

        # Sales insights and charts for the rep workbook.
        # Uses real order data from the uploaded commission report.
        all_rep_source_rows = []
        for level_rows in rep_level_rows[rep_name].values():
            all_rep_source_rows.extend(level_rows)

        if report_norm(rep_name) == "alex bethel":
            corrected_rows = []
            for source_row in all_rep_source_rows:
                source_row = source_row[:]
                source_row = source_row + [""] * (len(headers) - len(source_row))
                if not is_cancelled_order(source_row) and report_norm(source_row[COMMISSION_LEVELS[1]["name"]]) == "alex bethel":
                    source_row[COMMISSION_LEVELS[1]["pct"]] = "20%"
                    source_row[COMMISSION_LEVELS[1]["comm"]] = round(report_money(source_row[SUBTOTAL_IDX]) * 0.20, 2)
                corrected_rows.append(source_row)
            all_rep_source_rows = corrected_rows

        sales_metrics = add_sales_insight_tabs(
            wb,
            rep_name,
            all_rep_source_rows,
            headers,
            period,
            sales_history_df
        )

        history_update_rows.append(
            create_sales_history_row(rep_name, period, sales_metrics, generated_at)
        )

        file_name = f"{safe_filename(rep_name)} {period}.xlsx"
        file_path = os.path.join(rep_dir, file_name)
        wb.save(file_path)

        pay_entries.append({
            "Rep Name": rep_name,
            "Regular Commission": round(regular_due, 2),
            "Override": round(override_due, 2),
            "Total Due": round(total_due, 2),
            "Notes": override_note,
            "Report File": file_name,
            "Rows": row_count,
            "Cancelled": cancelled_count,
            "Orders": sales_metrics.get("orders", 0),
            "Sales": sales_metrics.get("sales", 0),
            "Average Order": sales_metrics.get("average_order", 0),
            "Top Products": ", ".join([p for p, c in sales_metrics.get("top_products", [])]),
            "Path": file_path,
        })

    # Joel override from Nelson
    nelson_rows = []
    for row in data:
        row = row + [""] * (len(headers) - len(row))
        if any("nelson" in report_norm(row[COMMISSION_LEVELS[level]["name"]]) for level in COMMISSION_LEVELS):
            nelson_rows.append(row)

    if nelson_rows:
        wb = Workbook()
        wb.remove(wb.active)
        q_rev, due = add_override_detail_sheet(
            wb,
            "Joel Override Report",
            f"Joel 2.5% Nelson Override {period}",
            headers,
            nelson_rows,
            JOEL_NELSON_RATE
        )
        file_name = f"Joel Override {period}.xlsx"
        file_path = os.path.join(rep_dir, file_name)
        wb.save(file_path)
        pay_entries.append({
            "Rep Name": "Joel",
            "Regular Commission": 0,
            "Override": round(due, 2),
            "Total Due": round(due, 2),
            "Notes": "2.5% of Nelson qualifying subtotal revenue",
            "Report File": file_name,
            "Rows": len(nelson_rows),
            "Cancelled": sum(1 for r in nelson_rows if is_cancelled_order(r)),
            "Path": file_path,
        })

    # Howard finder fee
    howard_rows = []
    for row in data:
        row = row + [""] * (len(headers) - len(row))
        email = report_norm(row[CUSTOMER_EMAIL_IDX])
        if email in HOWARD_CLIENT_EMAILS:
            howard_rows.append(row)

    if howard_rows:
        wb = Workbook()
        wb.remove(wb.active)
        q_rev, due = add_override_detail_sheet(
            wb,
            "Howard Finder Fee",
            f"Howard 5% Finder Fee {period}",
            headers,
            howard_rows,
            HOWARD_FINDER_RATE
        )
        file_name = f"Howard Finder Fee {period}.xlsx"
        file_path = os.path.join(rep_dir, file_name)
        wb.save(file_path)
        pay_entries.append({
            "Rep Name": "Howard",
            "Regular Commission": 0,
            "Override": round(due, 2),
            "Total Due": round(due, 2),
            "Notes": "5% finder fee from assigned client emails",
            "Report File": file_name,
            "Rows": len(howard_rows),
            "Cancelled": sum(1 for r in howard_rows if is_cancelled_order(r)),
            "Path": file_path,
        })

    # Master pay report
    pay_wb = Workbook()
    pay_ws = pay_wb.active
    pay_ws.title = "Master Pay Report"
    pay_ws["A1"] = f"Nu Life Commission Pay Report - {period}"
    pay_ws["A1"].font = Font(size=16, bold=True)
    pay_headers = ["Rep Name", "Regular Commission", "Override", "Total Due", "Notes", "Report File"]
    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="D9EAF7")

    for c, h in enumerate(pay_headers, 1):
        cell = pay_ws.cell(3, c, h)
        cell.fill = blue
        cell.font = Font(bold=True, color="FFFFFF")

    for rr, entry in enumerate(sorted(pay_entries, key=lambda x: x["Rep Name"].lower()), 4):
        for cc, key in enumerate(pay_headers, 1):
            pay_ws.cell(rr, cc, entry.get(key, ""))

    last_row = len(pay_entries) + 3
    total_row = last_row + 2
    pay_ws[f"C{total_row}"] = "Total Pay Due"
    pay_ws[f"D{total_row}"] = f"=SUM(D4:D{last_row})"
    pay_ws[f"C{total_row}"].font = Font(bold=True)
    pay_ws[f"D{total_row}"].font = Font(bold=True)
    pay_ws[f"C{total_row}"].fill = light
    pay_ws[f"D{total_row}"].fill = light

    for col in ["B", "C", "D"]:
        for cell in pay_ws[col][3:total_row]:
            cell.number_format = "$#,##0.00"

    apply_widths(pay_ws)
    master_pay_path = os.path.join(package_dir, "Nu Life Master Pay Report.xlsx")
    pay_wb.save(master_pay_path)

    # Update permanent rep_sales_history tab and create analytics workbook.
    updated_sales_history_df = upsert_sales_history(sales_history_df, history_update_rows)
    save_sales_history(updated_sales_history_df)

    analytics_wb = create_sales_analytics_workbook(updated_sales_history_df, period)
    analytics_path = os.path.join(package_dir, "Nu Life Sales Analytics.xlsx")
    analytics_wb.save(analytics_path)

    # Master all-in-one workbook
    master_wb = Workbook()
    master_wb.remove(master_wb.active)

    for source_file in [master_pay_path, analytics_path] + [e["Path"] for e in pay_entries]:
        src_wb = load_workbook(source_file, data_only=False)
        for sheet_name in src_wb.sheetnames:
            src_ws = src_wb[sheet_name]
            base_name = sheet_name if source_file == master_pay_path else f"{os.path.basename(source_file)[:15]} {sheet_name[:12]}"
            dest_name = re.sub(r'[:\\/*?\[\]]', '', base_name)[:31]
            original = dest_name
            counter = 1
            while dest_name in master_wb.sheetnames:
                suffix = f" {counter}"
                dest_name = (original[:31-len(suffix)] + suffix)
                counter += 1
            dst_ws = master_wb.create_sheet(dest_name)
            for row in src_ws.iter_rows():
                for cell in row:
                    new_cell = dst_ws.cell(cell.row, cell.column, cell.value)
                    if cell.has_style:
                        new_cell._style = copy(cell._style)
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = copy(cell.alignment)
            for col_key, dim in src_ws.column_dimensions.items():
                dst_ws.column_dimensions[col_key].width = dim.width
            dst_ws.freeze_panes = src_ws.freeze_panes

    master_all_path = os.path.join(package_dir, f"Nu Life Essentials {period}.xlsx")
    master_wb.save(master_all_path)

    # Email readiness lookup
    def lookup_email(rep_name):
        """
        Iron-clad matching order:
        1. Exact normalized FullName
        2. First + Last token match
        3. Last name + first initial
        4. Last name only when unique
        5. Fuzzy match above 0.86 when unique
        """

        rep_name_clean = normalize_name(rep_name)
        rep_parts = name_tokens(rep_name)
        rep_first = rep_parts[0] if rep_parts else ""
        rep_last = rep_parts[-1] if rep_parts else ""

        profiles = reps_df.copy()

        for required_col in ["FullName", "FirstName", "LastName", "NuLifeEmail", "PersonalEmail"]:
            if required_col not in profiles.columns:
                profiles[required_col] = ""

        profiles["FullName_clean"] = profiles["FullName"].apply(normalize_name)
        profiles["FirstName_clean"] = profiles["FirstName"].apply(normalize_name)
        profiles["LastName_clean"] = profiles["LastName"].apply(normalize_name)

        # Build a clean name from FirstName + LastName too, because FullName can be missing/stale.
        profiles["BuiltName_clean"] = (
            profiles["FirstName_clean"].astype(str).str.strip()
            + " "
            + profiles["LastName_clean"].astype(str).str.strip()
        ).str.strip()

        match = profiles[
            (profiles["FullName_clean"] == rep_name_clean)
            | (profiles["BuiltName_clean"] == rep_name_clean)
        ]

        match_method = "Exact name match"

        # First + last token match
        if match.empty and rep_first and rep_last:
            match = profiles[
                (profiles["FirstName_clean"] == rep_first)
                & (profiles["LastName_clean"] == rep_last)
            ]
            match_method = "First/last match"

        # Last name + first initial match
        if match.empty and rep_first and rep_last:
            match = profiles[
                (profiles["LastName_clean"] == rep_last)
                & (
                    profiles["FirstName_clean"]
                    .astype(str)
                    .str.startswith(rep_first[:1])
                )
            ]
            match_method = "Last name + first initial match"

        # Last name only, only if unique
        if match.empty and rep_last:
            last_matches = profiles[profiles["LastName_clean"] == rep_last]
            if len(last_matches) == 1:
                match = last_matches
                match_method = "Unique last-name match"

        # Fuzzy match as final fallback
        if match.empty:
            candidates = []
            for idx, row in profiles.iterrows():
                score_full = name_similarity(rep_name_clean, row.get("FullName_clean", ""))
                score_built = name_similarity(rep_name_clean, row.get("BuiltName_clean", ""))
                score = max(score_full, score_built)
                if score >= 0.86:
                    candidates.append((score, idx))

            candidates = sorted(candidates, reverse=True)

            if len(candidates) == 1:
                match = profiles.loc[[candidates[0][1]]]
                match_method = f"Fuzzy match {candidates[0][0]:.2f}"
            elif len(candidates) > 1:
                best_score = candidates[0][0]
                close = [c for c in candidates if best_score - c[0] < 0.03]
                if len(close) == 1:
                    match = profiles.loc[[close[0][1]]]
                    match_method = f"Fuzzy match {close[0][0]:.2f}"
                else:
                    possible = ", ".join(
                        profiles.loc[idx, "FullName"] for _, idx in close[:5]
                    )
                    return "", "", "", f"Multiple possible reps - {possible}"

        if match.empty:
            return "", "", "", "Rep not found"

        if len(match) > 1:
            possible = ", ".join(match["FullName"].astype(str).head(5).tolist())
            return "", "", "", f"Multiple possible reps - {possible}"

        row = match.iloc[0]

        nulife_email = report_clean(row.get("NuLifeEmail", ""))
        personal_email = report_clean(row.get("PersonalEmail", ""))

        if nulife_email:
            return nulife_email, nulife_email, personal_email, f"Ready - NuLifeEmail ({match_method})"

        if personal_email:
            return personal_email, nulife_email, personal_email, f"Ready - PersonalEmail ({match_method})"

        return "", nulife_email, personal_email, f"Missing email ({match_method})"

    for entry in sorted(pay_entries, key=lambda x: x["Rep Name"].lower()):
        send_to, nulife_email, personal_email, status = lookup_email(entry["Rep Name"])
        delivery_rows.append({
            "Rep Name": entry["Rep Name"],
            "NuLifeEmail": nulife_email,
            "PersonalEmail": personal_email,
            "Send To": send_to,
            "Status": status,
            "Report File": entry["Report File"],
            "Path": entry["Path"],
        })

    delivery_df = pd.DataFrame(delivery_rows)
    delivery_csv_path = os.path.join(package_dir, "email_delivery_readiness.csv")
    delivery_df.drop(columns=["Path"]).to_csv(delivery_csv_path, index=False)

    # Zip package
    zip_path = os.path.join(temp_root, f"{safe_filename(folder_name)}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for root, dirs, files in os.walk(package_dir):
            for file in files:
                local_path = os.path.join(root, file)
                arcname = os.path.relpath(local_path, temp_root)
                z.write(local_path, arcname)

    # Upload to Drive
    # Production approach: upload ONE ZIP file only.
    # This avoids Google Drive API failures from uploading many individual Excel files.
    drive_folder_link = ""
    drive_files_uploaded = []

    if REPORTS_PARENT_FOLDER_ID:
        try:
            service = get_drive_service()

            zip_drive_name = f"{safe_filename(folder_name)}.zip"
            uploaded_zip = upload_file_to_drive(
                service,
                zip_path,
                REPORTS_PARENT_FOLDER_ID,
                drive_name=zip_drive_name
            )

            drive_folder_link = uploaded_zip.get("webViewLink", "")
            drive_files_uploaded.append(drive_folder_link)

        except Exception as e:
            # Do not crash the reporting engine if Drive upload fails.
            # The ZIP is still available for download and emails can still send.
            drive_folder_link = ""
            drive_files_uploaded.append(f"Drive upload failed: {type(e).__name__}: {str(e)}")

    # Email send or test preview
    email_log = []
    if send_live:
        if not SENDER_EMAIL or not SENDER_APP_PASSWORD:
            raise RuntimeError("Missing SENDER_EMAIL or SENDER_APP_PASSWORD in Streamlit secrets.")

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_APP_PASSWORD)

        for item in delivery_rows:
            if not item["Send To"]:
                email_log.append({**item, "Email Status": "SKIPPED - Missing email"})
                continue

            to_email = test_email.strip() if test_email.strip() else item["Send To"]

            msg = EmailMessage()
            msg["Subject"] = folder_name
            msg["From"] = SENDER_EMAIL
            msg["To"] = to_email
            msg.set_content(
                f"Hello {item['Rep Name']},\n\n"
                f"Attached is your Nu Life commission report for this period.\n\n"
                f"Please review your report carefully.\n\n"
                f"Thank you,\nNu Life Essentials\n"
            )

            with open(item["Path"], "rb") as f:
                msg.add_attachment(
                    f.read(),
                    maintype="application",
                    subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    filename=item["Report File"]
                )

            server.send_message(msg)
            email_log.append({**item, "Email Status": f"SENT to {to_email}"})

        server.quit()
    else:
        for item in delivery_rows:
            if not item["Send To"]:
                email_log.append({**item, "Email Status": "TEST - Missing email"})
            else:
                email_log.append({**item, "Email Status": f"TEST - Would send to {item['Send To']}"})

    email_log_df = pd.DataFrame(email_log).drop(columns=["Path"])
    email_log_path = os.path.join(package_dir, "email_delivery_log.xlsx")
    email_log_df.to_excel(email_log_path, index=False)

    return {
        "folder_name": folder_name,
        "zip_path": zip_path,
        "package_dir": package_dir,
        "drive_folder_link": drive_folder_link,
        "drive_files_uploaded": drive_files_uploaded,
        "pay_entries": pd.DataFrame(pay_entries).drop(columns=["Path"]),
        "delivery_df": delivery_df.drop(columns=["Path"]),
        "email_log_df": email_log_df,
        "reports_count": len(pay_entries),
        "sales_history_df": updated_sales_history_df,
    }

def render_reporting_page(reps_df):
    st.title("Reporting")
    st.caption("Upload Nu Life commission CSV, generate reports, save to Drive, and email reps.")

    missing = []
    for key in ["REPORTS_PARENT_FOLDER_ID", "SENDER_EMAIL", "SENDER_APP_PASSWORD"]:
        if not st.secrets.get(key, ""):
            missing.append(key)

    if missing:
        st.warning("Missing Streamlit secrets: " + ", ".join(missing))

    uploaded = st.file_uploader("Upload raw Nu Life commission CSV", type=["csv"])

    c1, c2, c3 = st.columns(3)
    with c1:
        test_mode = st.checkbox("TEST MODE - do not send live emails", value=True)
    with c2:
        test_email = st.text_input("Optional test email override")
    with c3:
        confirm_live = st.text_input("Type SEND to allow live email", value="")

    st.info("Email priority: NuLifeEmail first. If blank, PersonalEmail. If both blank, report is flagged and skipped.")

    if uploaded:
        st.success(f"Loaded file: {uploaded.name}")

    if st.button("Generate Reports", type="primary", use_container_width=True, disabled=uploaded is None):
        send_live = (not test_mode) and confirm_live.strip().upper() == "SEND"

        if not test_mode and confirm_live.strip().upper() != "SEND":
            st.error("Live mode requires typing SEND.")
            st.stop()

        with st.spinner("Generating reports..."):
            sales_history_df = load_sales_history()

            result = build_commission_package(
                uploaded,
                reps_df,
                sales_history_df=sales_history_df,
                send_live=send_live,
                test_email=test_email
            )

        st.success(f"Generated {result['reports_count']} report(s).")

        if result["drive_folder_link"]:
            st.success("ZIP uploaded to Google Drive.")
            st.markdown(f"**Drive ZIP:** {result['drive_folder_link']}")
        else:
            st.warning("ZIP was generated, but it was not uploaded to Google Drive. Use the download button below.")
            if result.get("drive_files_uploaded"):
                st.caption(str(result.get("drive_files_uploaded")[-1]))

        st.subheader("Master Pay Preview")
        st.dataframe(result["pay_entries"], use_container_width=True)

        st.subheader("Sales Analytics Preview")
        leaderboard, fastest_growth, cancelled = build_analytics_tables(
            result["sales_history_df"],
            result["folder_name"]
        )

        tab1, tab2, tab3 = st.tabs(["Leaderboard", "Fastest Growing", "Cancelled %"])
        with tab1:
            st.dataframe(leaderboard, use_container_width=True)
        with tab2:
            st.dataframe(fastest_growth, use_container_width=True)
        with tab3:
            st.dataframe(cancelled, use_container_width=True)

        st.subheader("Email Delivery Preview / Log")
        st.dataframe(result["email_log_df"], use_container_width=True)

        email_log_file = os.path.join(result["package_dir"], "email_delivery_log.xlsx")
        if os.path.exists(email_log_file):
            with open(email_log_file, "rb") as f:
                st.download_button(
                    "Download Email Delivery Log",
                    data=f,
                    file_name="email_delivery_log.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        with open(result["zip_path"], "rb") as f:
            st.download_button(
                "Download Full Report Package ZIP",
                data=f,
                file_name=os.path.basename(result["zip_path"]),
                mime="application/zip",
                use_container_width=True
            )


def login():
    st.title("Nu Life Admin App")
    st.caption("Secure access required")

    pw = st.text_input("Password", type="password")

    if st.button("Login"):
        if pw == APP_PASSWORD:
            st.session_state.auth = True
            st.rerun()
        else:
            st.error("Wrong password")

if "auth" not in st.session_state:
    st.session_state.auth = False

if not st.session_state.auth:
    login()
    st.stop()

reps_df = load_reps()
sales_df = clean_sales_df(load_sales())

if os.path.exists("logo.png"):
    st.sidebar.image("logo.png", width=120)

st.sidebar.title("Nu Life Admin App")

page = st.sidebar.radio(
    "Navigation",
    ["Dashboard", "Map", "Rep Directory", "Sales Dashboard", "Reporting", "Manage Reps"]
)

if st.sidebar.button("Log out"):
    st.session_state.auth = False
    st.rerun()

if st.sidebar.button("Refresh Data"):
    st.cache_data.clear()
    st.rerun()

if page == "Dashboard":
    st.title("Dashboard")

    working_df = reps_df.copy()
    working_df["Latitude"] = pd.to_numeric(working_df["Latitude"], errors="coerce")
    working_df["Longitude"] = pd.to_numeric(working_df["Longitude"], errors="coerce")

    active_df = working_df[working_df["Active"].astype(str).str.lower() == "yes"]

    total_revenue = sales_df["Revenue"].sum() if not sales_df.empty else 0
    total_orders = sales_df["Orders"].sum() if not sales_df.empty else 0

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Reps", len(working_df))
    c2.metric("Active Reps", len(active_df))
    c3.metric("Markets", working_df["MarketTerritory"].replace("", pd.NA).dropna().nunique())
    c4.metric("Total Revenue", f"${total_revenue:,.0f}")
    c5.metric("Total Orders", f"{int(total_orders):,}")

    st.markdown("---")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Reps by Manager")
        st.bar_chart(working_df["Manager"].replace("", "Unassigned").value_counts())

    with col2:
        st.subheader("Revenue by Rep")
        if not sales_df.empty:
            st.bar_chart(sales_df.groupby("FullName")["Revenue"].sum().sort_values(ascending=False))
        else:
            st.info("No sales data yet.")

    st.markdown("---")

    missing_coords = working_df[
        working_df["Latitude"].isna() | working_df["Longitude"].isna()
    ]

    st.subheader("Data Alerts")
    if missing_coords.empty:
        st.success("All reps have map coordinates.")
    else:
        st.warning(f"{len(missing_coords)} rep(s) are missing Latitude/Longitude.")
        st.dataframe(missing_coords, use_container_width=True)

elif page == "Map":
    st.title("NuLife Rep Map")

    map_df = reps_df.copy()
    map_df["Latitude"] = pd.to_numeric(map_df["Latitude"], errors="coerce")
    map_df["Longitude"] = pd.to_numeric(map_df["Longitude"], errors="coerce")
    map_df = map_df.dropna(subset=["Latitude", "Longitude"]).reset_index(drop=True)

    st.subheader("Filters")

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        states = ["All"] + sorted(map_df["State"].dropna().astype(str).unique().tolist())
        selected_state = st.selectbox("State", states)

    with col2:
        managers = ["All"] + sorted(map_df["Manager"].dropna().astype(str).unique().tolist())
        selected_manager = st.selectbox("Manager", managers)

    with col3:
        regions = ["All"] + sorted(map_df["Region"].dropna().astype(str).unique().tolist())
        selected_region = st.selectbox("Region", regions)

    with col4:
        search = st.text_input("Search")

    filtered_df = map_df.copy()

    if selected_state != "All":
        filtered_df = filtered_df[filtered_df["State"].astype(str) == selected_state]

    if selected_manager != "All":
        filtered_df = filtered_df[filtered_df["Manager"].astype(str) == selected_manager]

    if selected_region != "All":
        filtered_df = filtered_df[filtered_df["Region"].astype(str) == selected_region]

    if search:
        mask = filtered_df.astype(str).apply(
            lambda row: row.str.contains(search, case=False, na=False).any(),
            axis=1
        )
        filtered_df = filtered_df[mask]

    filtered_df = filtered_df.reset_index(drop=True)

    st.markdown(f"### Showing {len(filtered_df)} Rep(s)")

    m = folium.Map(location=[39.5, -98.35], zoom_start=4, tiles="OpenStreetMap")

    for i, (_, row) in enumerate(filtered_df.iterrows()):
        offset_lat, offset_lng = stable_offset(i)
        lat = float(row["Latitude"]) + offset_lat
        lng = float(row["Longitude"]) + offset_lng

        rep_sales = sales_df[sales_df["RepID"].astype(str) == str(row.get("RepID", ""))]
        rep_revenue = rep_sales["Revenue"].sum() if not rep_sales.empty else 0
        rep_orders = rep_sales["Orders"].sum() if not rep_sales.empty else 0

        popup_html = f"""
        <div style="width:280px; font-family: Arial, sans-serif;">
            <h4>{row.get('FullName', '')}</h4>
            <b>Rep ID:</b> {row.get('RepID', '')}<br>
            <b>Territory:</b> {row.get('MarketTerritory', '')}<br>
            <b>City/State:</b> {row.get('City', '')}, {row.get('State', '')}<br>
            <b>Manager:</b> {row.get('Manager', '')}<br>
            <b>Region:</b> {row.get('Region', '')}<br><br>
            <b>Total Revenue:</b> ${rep_revenue:,.0f}<br>
            <b>Total Orders:</b> {int(rep_orders)}<br><br>
            <b>Phone:</b><br>{row.get('PhoneNumber', '')}<br><br>
            <b>Email:</b><br>{row.get('PersonalEmail', '')}<br><br>
            <b>NuLife Email:</b><br>{row.get('NuLifeEmail', '')}<br><br>
            <b>Business:</b><br>{row.get('BusinessName', '')}<br><br>
            <b>Notes:</b><br>{row.get('Notes', '')}
        </div>
        """

        folium.Marker(
            [lat, lng],
            popup=folium.Popup(popup_html, max_width=340),
            tooltip=row.get("FullName", "Rep"),
            icon=folium.Icon(color="blue", icon="flag")
        ).add_to(m)

    st_folium(m, width=1150, height=650, returned_objects=[], key="rep_map")

elif page == "Rep Directory":
    st.title("Rep Directory")

    search_dir = st.text_input("Search reps, markets, managers, states")
    directory_df = reps_df.copy()

    if search_dir:
        mask = directory_df.astype(str).apply(
            lambda row: row.str.contains(search_dir, case=False, na=False).any(),
            axis=1
        )
        directory_df = directory_df[mask]

    st.markdown(f"### {len(directory_df)} Rep(s)")

    for _, row in directory_df.iterrows():
        rep_sales = sales_df[sales_df["RepID"].astype(str) == str(row.get("RepID", ""))]
        rep_revenue = rep_sales["Revenue"].sum() if not rep_sales.empty else 0
        rep_orders = rep_sales["Orders"].sum() if not rep_sales.empty else 0

        st.markdown(
            f"""
            <div class="premium-card">
                <div class="premium-card-title">{row.get('FullName', '')}</div>
                <div class="premium-card-subtitle">
                    {row.get('MarketTerritory', '')} • {row.get('City', '')}, {row.get('State', '')} • Manager: {row.get('Manager', '')}
                </div>
                <div class="premium-card-body">
                    <b>Revenue:</b> ${rep_revenue:,.0f} &nbsp; | &nbsp;
                    <b>Orders:</b> {int(rep_orders)}<br><br>
                    <b>Phone:</b> {row.get('PhoneNumber', '')}<br>
                    <b>Email:</b> {row.get('PersonalEmail', '')}<br>
                    <b>NuLife:</b> {row.get('NuLifeEmail', '')}<br><br>
                    {row.get('Notes', '')}
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

elif page == "Sales Dashboard":
    st.title("Sales Dashboard")

    sales_history_df = load_sales_history()

    if not sales_history_df.empty:
        st.subheader("Real Sales History Analytics")

        history_df = normalize_history_df(sales_history_df)
        periods = sorted(history_df["ReportPeriod"].dropna().astype(str).unique().tolist(), key=period_sort_key)

        if periods:
            selected_period = st.selectbox("Report Period", periods, index=len(periods) - 1)

            leaderboard, fastest_growth, cancelled = build_analytics_tables(history_df, selected_period)

            tab1, tab2, tab3 = st.tabs(["Leaderboard", "Fastest Growing", "Cancelled %"])
            with tab1:
                st.dataframe(leaderboard, use_container_width=True)
                if not leaderboard.empty:
                    st.bar_chart(leaderboard.set_index("RepName")["Sales"])
            with tab2:
                st.dataframe(fastest_growth, use_container_width=True)
            with tab3:
                st.dataframe(cancelled, use_container_width=True)

            st.markdown("---")

    if sales_df.empty:
        if sales_history_df.empty:
            st.warning("No sales data found yet. Run a commission report to populate rep_sales_history.")
        st.stop()

    total_revenue = sales_df["Revenue"].sum()
    total_orders = sales_df["Orders"].sum()
    total_providers = sales_df["Providers"].sum()
    avg_order_value = total_revenue / total_orders if total_orders else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total Revenue", f"${total_revenue:,.0f}")
    c2.metric("Total Orders", f"{int(total_orders):,}")
    c3.metric("Providers", f"{int(total_providers):,}")
    c4.metric("Avg Order Value", f"${avg_order_value:,.0f}")

    st.markdown("---")

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Rep Leaderboard")
        leaderboard = sales_df.groupby("FullName", as_index=False).agg({
            "Revenue": "sum",
            "Orders": "sum",
            "Providers": "sum"
        }).sort_values("Revenue", ascending=False)
        st.dataframe(leaderboard, use_container_width=True)

    with col2:
        st.subheader("Revenue by Territory")
        st.bar_chart(sales_df.groupby("MarketTerritory")["Revenue"].sum().sort_values(ascending=False))

    st.markdown("---")

    col3, col4 = st.columns(2)

    with col3:
        st.subheader("Orders by Rep")
        st.bar_chart(sales_df.groupby("FullName")["Orders"].sum().sort_values(ascending=False))

    with col4:
        st.subheader("Top Products")
        st.bar_chart(sales_df.groupby("TopProduct")["Revenue"].sum().sort_values(ascending=False))

    st.markdown("---")
    st.subheader("Raw Sales Data")
    st.dataframe(sales_df, use_container_width=True)

elif page == "Reporting":
    render_reporting_page(reps_df)

elif page == "Manage Reps":
    st.title("Manage Reps")

    st.subheader("Add New Rep")

    with st.form("add_rep_form"):
        c1, c2, c3 = st.columns(3)

        with c1:
            first_name = st.text_input("First Name")
            last_name = st.text_input("Last Name")
            active = st.selectbox("Active", ["Yes", "No"], index=0)
            manager = st.text_input("Manager")

        with c2:
            region = st.text_input("Region")
            market = st.text_input("Market / Territory")
            state = st.text_input("State")
            city = st.text_input("City")

        with c3:
            phone = st.text_input("Phone Number")
            personal_email = st.text_input("Personal Email")
            nulife_email = st.text_input("NuLife Email")
            links = st.text_input("Links / Handles")

        business = st.text_input("Business Name")
        address = st.text_input("Address")

        c4, c5 = st.columns(2)
        with c4:
            latitude = st.text_input("Latitude")
        with c5:
            longitude = st.text_input("Longitude")

        notes = st.text_area("Notes")
        submitted = st.form_submit_button("Add Rep")

        if submitted:
            if not first_name.strip() or not last_name.strip():
                st.error("First Name and Last Name are required.")
            else:
                new_rep_id = generate_next_rep_id(reps_df)
                full_name = f"{first_name.strip()} {last_name.strip()}"
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                new_row = {
                    "RepID": new_rep_id,
                    "Active": active,
                    "Manager": manager,
                    "Region": region,
                    "MarketTerritory": market,
                    "State": state,
                    "City": city,
                    "FirstName": first_name,
                    "LastName": last_name,
                    "FullName": full_name,
                    "PhoneNumber": phone,
                    "PersonalEmail": personal_email,
                    "NuLifeEmail": nulife_email,
                    "LinksHandles": links,
                    "BusinessName": business,
                    "Address": address,
                    "Latitude": latitude,
                    "Longitude": longitude,
                    "Notes": notes,
                    "StartDate": now,
                    "LastUpdated": now
                }

                updated_df = pd.concat([reps_df, pd.DataFrame([new_row])], ignore_index=True)

                if save_reps(updated_df):
                    st.success(f"Added {full_name} as {new_rep_id}.")
                    st.rerun()

    st.markdown("---")

    st.subheader("Edit Existing Reps")
    st.info("Edit reps below, then click Save Changes to update Google Sheets.")

    edited_df = st.data_editor(
        reps_df.copy(),
        use_container_width=True,
        num_rows="dynamic",
        key="rep_editor"
    )

    c1, c2 = st.columns(2)

    with c1:
        if st.button("Save Changes", type="primary", use_container_width=True):
            edited_df["LastUpdated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            if save_reps(edited_df):
                st.success("Rep profiles saved successfully.")
                st.rerun()

    with c2:
        if st.button("Discard Changes / Refresh", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

   
